#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 10 21:47:04 2018

@author: jayhan
"""
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")

## access cells
for ws in wb.worksheets:
    for line in ws:
        for cell in line:
            print(cell.value)

##get cell value
ws_names = wb.sheetnames
ws = wb[ws_names[1]]
cell = ws['C2']
print(cell.value)
cell = ws.cell(column=10,row=3)
print(cell.value)

##get table
ws = wb[ws_names[0]]
table= ws.values
cols = next(table)
print(cols)

data = list(table)
print(data)


            
            
