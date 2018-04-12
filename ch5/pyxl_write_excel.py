#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 10 16:15:27 2018

@author: jayhan
"""

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
ws.cell(column=1,row=3, value='A2')

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()


# Rows can also be appended
ws.append([1, 2, 3])


ws = wb.create_sheet('2ndsheet')

# now we'll fill it with 100 rows x 200 columns
for irow in range(100):
    ws.append(['%d' % i for i in range(200)])

## switch ws
print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]
ws = wb.worksheets[0]
ws.append([1, 2, 3])
    
# Save the file
wb.save("sample.xlsx")